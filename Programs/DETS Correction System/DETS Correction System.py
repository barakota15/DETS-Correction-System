import cv2
import numpy as np
import pytesseract
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import time
import os
import shutil
from PIL import Image

# important: Change the path to tesseract path
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

num_questions = None
num_questions_int = None
# Load the template
template_image_path = None
def template():
    global num_questions
    global template_image_path
    global num_questions_int

    while True:
        num_questions = input("Select Template (10 - 30 - 60): ")
        if num_questions == "10":
            template_image_path = "Data/Templates/Template-10.png"
            num_questions_int = 10
            return True
        elif num_questions == "30":
            template_image_path = "Data/Templates/Template-30.png"
            num_questions_int = 30
            return True
        elif num_questions == "60":
            template_image_path = "Data/Templates/Template-60.png"
            num_questions_int = 60
            return True
        else:
            print("Invalid input. Please enter (10 - 30 - 60)")

template()

folder_path = input("Folder Path: ")
output_directory = "Data/output"
if os.path.isdir(output_directory):
    shutil.rmtree(output_directory)
os.makedirs(output_directory)
result_directory = "Data/results"
if os.path.isdir(result_directory):
    shutil.rmtree(result_directory)
os.makedirs(result_directory)
Excel_path = "Data/Excels"
if os.path.isdir(Excel_path):
    shutil.rmtree(Excel_path)
os.makedirs(Excel_path)

# Load Students Data
Students_data_path = input("Students Data Excel Sheet Path (Press y if it on the default path): ")
if Students_data_path == "y":
    Students_data_path = "Serial Number Maker/Students Data.xlsx"
print(f"Students Data Excel Sheet Path is: {Students_data_path}")

ANSWER_KEY = {}
correcte_answer = []
corrected_tests = {}
test_codes = {}
test_serial_codes = {}
serial_numbers = {}

def calculate_time(start_time, end_time):
    elapsed_time = end_time - start_time
    return elapsed_time

def format_time(seconds):
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

def align_image(template_image_path, input_image_path, output_aligned_image_path):
    img = Image.open(input_image_path)
    img = img.convert('L')
    threshold_img = img.point(lambda p: p > 220 and 255)
    threshold_img.save(output_aligned_image_path)

    template = cv2.imread(template_image_path, 0)
    input_image = cv2.imread(output_aligned_image_path, 0)

    # Initiate the SIFT detector
    sift = cv2.SIFT_create()

    # Find the keypoints and descriptors with SIFT
    kp1, des1 = sift.detectAndCompute(template, None)
    kp2, des2 = sift.detectAndCompute(input_image, None)

    # FLANN parameters
    FLANN_INDEX_KDTREE = 1
    index_params = dict(algorithm=FLANN_INDEX_KDTREE, trees=5)
    search_params = dict(checks=50)

    flann = cv2.FlannBasedMatcher(index_params, search_params)
    matches = flann.knnMatch(des1, des2, k=2)

    # Ratio test to filter good matches
    good_matches = []
    for m, n in matches:
        if m.distance < 0.7 * n.distance:
            good_matches.append(m)

    if len(good_matches) > 4:
        src_pts = np.float32([kp1[m.queryIdx].pt for m in good_matches]).reshape(-1, 1, 2)
        dst_pts = np.float32([kp2[m.trainIdx].pt for m in good_matches]).reshape(-1, 1, 2)

        # Calculate the transformation matrix
        M, _ = cv2.findHomography(dst_pts, src_pts, cv2.RANSAC, 5.0)

        # Apply the transformation to the input image
        result = cv2.warpPerspective(input_image, M, (template.shape[1], template.shape[0]))

        # Save the aligned image
        cv2.imwrite(output_aligned_image_path, result)

        print(f"Aligned image saved at {output_aligned_image_path}")
    else:
        print("Not enough matches are found - %d/%d" % (len(good_matches), 4))
        
    Scan(output_aligned_image_path)
    
def encode(serial_number):
    test_code = ""
    for char in serial_number:
        test_code += str(ord(char))
    return test_code
    
def extract_serial_number(image):
    # Convert the image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Apply image preprocessing techniques
    processed_image = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 85, 11)

    # Use pytesseract to extract text
    extracted_text = pytesseract.image_to_string(processed_image)

    # Search for the serial number pattern
    serial_number = None
    for line in extracted_text.split('\n'):
        if 'S/N: ' in line:
            serial_number = line.split('S/N: ')[1]
            break

    return serial_number

def is_shaded(img, x, y, radius, threshold):
    height, width, _ = img.shape
    total_color = np.array([0, 0, 0], dtype=np.uint64)
    count = 0

    for j in range(-radius, radius + 1):
        for i in range(-radius, radius + 1):
            if 0 <= y + j < height and 0 <= x + i < width:
                total_color += img[y + j, x + i]
                count += 1

    if count > 0:
        average_color = total_color / count
        return sum(average_color) < threshold * 3
    else:
        return False

def Scan(image_path):
    # Load the image
    img = cv2.imread(image_path)

    # Define the radius and shading threshold
    radius = 10
    shaded_threshold = 190

    # Draw a point on the image at the specified coordinates
    color = (0, 0, 255)  # Red color in BGR
    thickness = -1  # Fill the circle
    radius_circle = 15  # Radius of the circle
    
    Student_Answers = {} # Answers of students
    
    if num_questions_int == 10:
        # Define the coordinates
        x, y = 306, 653

        # Define the radius and shading threshold
        radius = 20
        shaded_threshold = 150
        for j in range(num_questions_int):
            shaded = 0
            for i in range(4):
                # Calculate the coordinates
                point_x = x + 120 * i
                point_y = y + 120 * j

                # Check if the calculated coordinates are within the image dimensions
                if point_x < img.shape[1] and point_y < img.shape[0]:
                    # Check if the area is shaded
                    if is_shaded(img, point_x, point_y, radius, shaded_threshold):
                        Student_Answers[j] = i
                        cv2.circle(img, (point_x, point_y), radius_circle, color, thickness)
                        shaded += 1
                    
                    if shaded > 1:
                        Student_Answers[j] = 4
                    
                    if i == 3 and j not in Student_Answers:
                        Student_Answers[j] = 5
    elif num_questions_int == 30:
        # Define the coordinates
        x, y = 353, 614

        # Define the radius and shading threshold
        radius = 20
        shaded_threshold = 170
        for j in range(num_questions_int):
            shaded = 0
            if j == 15:
                x, y = 1022, 614
            for i in range(4):
                # Calculate the coordinates
                point_x = x + 94 * i
                point_y = y + 85 * j
                if j > 14:
                    point_y = y + 85 * (j-15)

                # Check if the calculated coordinates are within the image dimensions
                if point_x < img.shape[1] and point_y < img.shape[0]:
                    # Check if the area is shaded
                    if is_shaded(img, point_x, point_y, radius, shaded_threshold):
                        Student_Answers[j] = i
                        cv2.circle(img, (point_x, point_y), radius_circle, color, thickness)
                        shaded += 1
                    
                    if shaded > 1:
                        Student_Answers[j] = 4
                    
                    if i == 3 and j not in Student_Answers:
                        Student_Answers[j] = 5
    elif num_questions_int == 60:
        # Define the coordinates
        x, y = 242, 620

        # Define the radius and shading threshold
        radius = 20
        shaded_threshold = 170
        for j in range(num_questions_int):
            shaded = 0
            if j == 20:
                x, y = 696, 620
            elif j == 40:
                x, y = 1159, 620
            for i in range(4):
                # Calculate the coordinates
                point_x = x + 84 * i
                point_y = y + 63 * j
                if j > 19 and j < 40:
                    point_y = y + 63 * (j-20)
                elif j > 39:
                    point_y = y + 63 * (j-40)

                # Check if the calculated coordinates are within the image dimensions
                if point_x < img.shape[1] and point_y < img.shape[0]:
                    # Check if the area is shaded
                    if is_shaded(img, point_x, point_y, radius, shaded_threshold):
                        Student_Answers[j] = i
                        cv2.circle(img, (point_x, point_y), radius_circle, color, thickness)
                        shaded += 1
                    
                    if shaded > 1:
                        Student_Answers[j] = 4
                    
                    if i == 3 and j not in Student_Answers:
                        Student_Answers[j] = 5

    Correction(Student_Answers, img)

def Correction(Student_Answers, image):
    #The number of correct answers for the student
    correct = 0
    
    #Answer List
    Answer_list = []
    
    for i in range(num_questions_int):
        if ANSWER_KEY[i] == Student_Answers[i]:
            correct += 1
        
        if Student_Answers[i] == 0:
            Answer_list.extend('A')
        elif Student_Answers[i] == 1:
            Answer_list.extend('B')
        elif Student_Answers[i] == 2:
            Answer_list.extend('C')
        elif Student_Answers[i] == 3:
            Answer_list.extend('D')
        elif Student_Answers[i] == 4:
            Answer_list.extend('M')
        elif Student_Answers[i] == 5:
            Answer_list.extend('N')
            
    serial_number = extract_serial_number(image)
    serial_numbers[os.path.basename(output_path)] = serial_number
    
    test_code = encode(serial_number)
    test_codes[serial_number] = test_code
    
    test_serial_codes[test_code] = serial_number
    
    # Draw a square on the image
    x, y, w, h = 122, 421, 339, 68
    cv2.rectangle(image, (x, y), (x + w, y + h), (0, 0, 0), -1)
    cv2.putText(image, test_code, (x, y + 115), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 0, 0), 2)
    
    # Save the aligned image with points
    result_path = f'{result_directory}/{test_code}.png'
    cv2.imwrite(result_path, image)
    corrected_tests[test_code] = {'answers': Answer_list, 'grade': correct, 'link': result_path}
    #print(corrected_tests[test_code])
    cv2.waitKey(0)
    
def create_excel_sheet(folder_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Results"
    green_count_formula = "0"
    
    # Writing headers
    sheet.cell(row=1, column=1, value="Test Code")
    questions = len(ANSWER_KEY)
    for i in range(questions):
        question_num = 'Question ' + str(i + 1)
        sheet.cell(row=1, column=i + 2, value=question_num)
    sheet.cell(row=1, column=questions + 2, value="Final Grade")
    sheet.cell(row=1, column=questions + 3, value="Link to Corrected Test")
    
    # Writing data
    row_num = 2
    for filename in os.listdir(folder_path):
        if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith('.jpeg'):
            image_path = os.path.join(folder_path, filename)
            test_name = os.path.basename(image_path)
            serial_number = serial_numbers[test_name]
            test_code = test_codes[serial_number]
            sheet.cell(row=row_num, column=1, value=test_code)
            for i, answer in enumerate(corrected_tests[test_code]['answers']):
                sheet.cell(row=row_num, column=i + 2, value=answer)
                counter = f'COUNTIF({get_column_letter(i + 2)}{row_num}, "{correcte_answer[i]}")'
                if i > 0:
                    green_count_formula += f"+{counter}"   
                else:
                    green_count_formula = f"={counter}"
            
            # Marking cells based on correctness
            grade_cell = f'{get_column_letter(questions + 2)}{row_num}'
            sheet[grade_cell] = green_count_formula
            
            # Adding hyperlink to the cell
            hyperlink_cell = sheet.cell(row=row_num, column=questions + 3)
            hyperlink_cell.value = f"Link to Corrected Test {test_code}"
            hyperlink_cell.font = Font(underline='single', color="0563C1")
            hyperlink = f"file:///{os.path.abspath(result_directory)}/{test_code}.png"
            sheet.cell(row=row_num, column=questions + 3).hyperlink = hyperlink
            sheet.cell(row=row_num, column=questions + 3).style = "Hyperlink"
            row_num += 1
        
    # Color the cells according to the answer
    for row in range(2, row_num):
        for col in range(2, questions + 2):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value == correcte_answer[col - 2]:
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Good theme
            elif cell_value == "N" or cell_value == "M":
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="FBFF7A", end_color="FBFF7A", fill_type="solid")  # Manually theme
            else:
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")  # Bad theme


    # Adding conditional formatting
    for col in range(2, questions + 2):
        rule = CellIsRule(operator="equal", formula=[f'"{correcte_answer[col - 2]}"'], stopIfTrue=True, fill=PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"))
        sheet.conditional_formatting.add(f'{sheet.cell(2, col).column_letter}2:{sheet.cell(row_num - 1, col).column_letter}{row_num - 1}', rule)

        rule = CellIsRule(operator="equal", formula=['"N"'], stopIfTrue=True, fill=PatternFill(start_color="FBFF7A", end_color="FBFF7A", fill_type="solid"))
        sheet.conditional_formatting.add(f'{sheet.cell(2, col).column_letter}2:{sheet.cell(row_num - 1, col).column_letter}{row_num - 1}', rule)

        rule = CellIsRule(operator="equal", formula=['"M"'], stopIfTrue=True, fill=PatternFill(start_color="FBFF7A", end_color="FBFF7A", fill_type="solid"))
        sheet.conditional_formatting.add(f'{sheet.cell(2, col).column_letter}2:{sheet.cell(row_num - 1, col).column_letter}{row_num - 1}', rule)

        rule = CellIsRule(operator="notEqual", formula=[f'"{correcte_answer[col - 2]}"', '"N"', '"M"'], stopIfTrue=True, fill=PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"))
        sheet.conditional_formatting.add(f'{sheet.cell(2, col).column_letter}2:{sheet.cell(row_num - 1, col).column_letter}{row_num - 1}', rule)

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        sheet.column_dimensions[f'{get_column_letter(questions + 2)}'].width = 11
            
    excel_file_path = os.path.join(Excel_path, f"{Excel_name}.xlsx")
    wb.save(excel_file_path)
    print(f"Excel file saved at {excel_file_path}")
    # Open the Excel file automatically
    os.startfile(excel_file_path)
    
def create_final_grades_excel(excel_file_path, test_serial_codes, Students_data_path):
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheet = wb.active

    final_grades_wb = openpyxl.Workbook()
    final_grades_sheet = final_grades_wb.active
    final_grades_sheet.title = "Final Grades"
    questions = len(ANSWER_KEY)
    
    # Writing headers
    final_grades_sheet.cell(row=1, column=1, value="Serial Number")
    final_grades_sheet.cell(row=1, column=2, value="Name")
    final_grades_sheet.cell(row=1, column=3, value="Acadmic Number")
    final_grades_sheet.cell(row=1, column=4, value="Final Grade")

    # Loading student data
    student_data_wb = openpyxl.load_workbook(Students_data_path)
    student_sheet = student_data_wb.active

    student_data = {}
    for row in student_sheet.iter_rows(min_row=2, max_row=student_sheet.max_row, max_col=3):
        serial_number = row[0].value
        name = row[1].value
        acadmic_number = row[2].value
        student_data[serial_number] = {'name': name, 'acadmic_number': acadmic_number}

    row_num = 2
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        test_code = row[0].value
        serial_number = test_serial_codes[test_code]
        final_grade = row[questions + 1].value
        if serial_number in student_data:
            name = student_data[serial_number]['name']
            acadmic_number = student_data[serial_number]['acadmic_number']
            final_grades_sheet.cell(row=row_num, column=1, value=serial_number)
            final_grades_sheet.cell(row=row_num, column=2, value=name)
            final_grades_sheet.cell(row=row_num, column=3, value=acadmic_number)
            final_grades_sheet.cell(row=row_num, column=4, value=final_grade)
            row_num += 1
        else:
            print(f"Serial number {serial_number} not found in the student data.")

    for column_cells in final_grades_sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        final_grades_sheet.column_dimensions[column_cells[0].column_letter].width = length + 2            
                    
    final_grades_file_path = "Final_Grades.xlsx"
    final_grades_wb.save(final_grades_file_path)
    
    # Protecting the sheet
    final_grades_wb = openpyxl.load_workbook(final_grades_file_path)
    final_grades_sheet = final_grades_wb.active
    final_grades_sheet.protection.sheet = True

    # Set a password for the sheet (optional)
    final_grades_sheet.protection.password = "securepassword"
    
    final_grades_wb.save(final_grades_file_path)
    print(f"Final grades Excel file saved at {final_grades_file_path}")    

i = 0    
while True:
    question_num = 'Question ' + str(i + 1)
    answer_letter = input(f"Enter the answer for {question_num} (A, B, C, D, G for None): ").upper()
    while answer_letter not in ['A', 'B', 'C', 'D', 'G']:
        answer_letter = input("Invalid input. Please enter A, B, C, D or G for None: ").upper()
            
    correcte_answer.extend(answer_letter)
    if answer_letter == 'A':
        ANSWER_KEY[i] = 0
    elif answer_letter == 'B':
        ANSWER_KEY[i] = 1
    elif answer_letter == 'C':
        ANSWER_KEY[i] = 2
    elif answer_letter == 'D':
        ANSWER_KEY[i] = 3
    elif answer_letter == 'G':
        ANSWER_KEY[i] = 6
            
    if i == num_questions_int - 1:
        print(correcte_answer)
        continuee = input("is the answers above right?! (y/n): ")
        if continuee in ['y', 'yes']:
            break
        elif continuee in ['n', 'no']:
            correcte_answer = []
            ANSWER_KEY = {}
            i = -1
            print("Alright, maybe next time.")
        else:
            correcte_answer = []
            ANSWER_KEY = {}
            i = -1
            print("Invalid input. Please enter 'y' for Yes or 'n' for No.") 
    i += 1            

start_time = time.time()      
for filename in os.listdir(folder_path):
    if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith('.jpeg'):
        image_path = os.path.join(folder_path, filename)
        output_path = os.path.join(output_directory, f"aligned_{filename}")
        align_image(template_image_path, image_path, output_path)

end_time = time.time()
elapsed_time = calculate_time(start_time, end_time)
formatted_time = format_time(elapsed_time)
print(f"Time elapsed: {formatted_time}")

now = datetime.now()
current_time = now.strftime("%H-%M")
Excel_name = f"Test_Grades_{current_time}"
create_excel_sheet(output_directory)

def submit(question):
    while True:
        answer = input(f"{question} (y/n): ").strip().lower()
        if answer in ['y', 'yes']:
            return True
        elif answer in ['n', 'no']:
            print("Alright, maybe next time.")
        else:
            print("Invalid input. Please enter 'y' for Yes or 'n' for No.")
            
respone = submit("Did you check the grades?")
if respone:
    create_final_grades_excel(os.path.join(Excel_path, f"{Excel_name}.xlsx"), test_serial_codes, Students_data_path)
    
def done(question):
    while True:
        answer = input(question)
        return True

done("press enter to close...")  