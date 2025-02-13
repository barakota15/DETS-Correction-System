import openpyxl
import random

# Load the workbook
wb = openpyxl.load_workbook('Students Data.xlsx')
sheet = wb.active

# Create a new workbook
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Write headers
new_sheet.cell(row=1, column=1, value="Serial Number")
new_sheet.cell(row=1, column=2, value="Name")
new_sheet.cell(row=1, column=3, value="Acadmic Number")

# Function to generate a random unique serial number
def generate_serial_number(existing_serial_numbers):
    while True:
        serial_number = f"BH{random.randint(100000, 999999)}"
        if serial_number not in existing_serial_numbers:
            return serial_number

# Store generated serial numbers
generated_serial_numbers = set()

# Iterate through each row in the original sheet
row_num = 2
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=2):
    name = row[0].value
    acadmic_number = row[1].value
    
    # Generate a unique serial number
    serial_number = generate_serial_number(generated_serial_numbers)
    generated_serial_numbers.add(serial_number)
    
    # Write the data to the new sheet
    new_sheet.cell(row=row_num, column=1, value=serial_number)
    new_sheet.cell(row=row_num, column=2, value=name)
    new_sheet.cell(row=row_num, column=3, value=acadmic_number)
    
    row_num += 1

# Save the new workbook
new_wb.save('Students Data.xlsx')
print("Excel file saved successfully.")

def done(question):
    while True:
        answer = input(question)
        return True

done("press enter to close...") 