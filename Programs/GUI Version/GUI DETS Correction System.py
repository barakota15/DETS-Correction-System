import cv2
import numpy as np
import pytesseract
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import time
import os
import shutil
from PIL import Image
import customtkinter
from customtkinter import filedialog
from tkinter import messagebox
import sys
import threading

# important: Change the path to tesseract path
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

num_questions = None

folder_path = "Click Browse To Choose Folder..."
Students_data_path = "Click Browse To Choose Excel File..."
output_path = None
Excel_name = None
final_grades_file_path = None

output_directory = "Data/output"
result_directory = "Data/results"
Excel_path = "Data/Excels"

# Load the template
template_image_path = None

ANSWER_KEY = {}
correcte_answer = []
corrected_tests = {}
test_codes = {}
test_serial_codes = {}
serial_numbers = {}

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")

app = customtkinter.CTk()
app.geometry("1000x800")

#app.resizable(False,False)
app.title("DETS Correction System")

MainFrame = customtkinter.CTkFrame(app)
MainFrame.pack(fill="both", expand=True)

Main = customtkinter.CTkFrame(MainFrame)
Main.pack(padx=30, pady=30, fill="both", expand=True)

def OpenFile():
    global Students_data_path
    filepath = filedialog.askopenfilename(title="Open Student Date Excel File", filetypes=(("Excel File","*.xlsx"),("All Files","*.*")))
    if filepath == "":
        messagebox.showinfo(title="Missing", message="You didn't Choose Excel File")
    else:
        Students_data_path = filepath
        Start()

def OpenFolder():
    global folder_path
    Folderpath = filedialog.askdirectory(title="Open Exams Folder")
    if Folderpath == "":
        messagebox.showinfo(title="Missing", message="You didn't Choose Folder")
    else:
        num_exams = 0
        for filename in os.listdir(Folderpath):
                if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith('.jpeg'):
                    num_exams += 1
        if num_exams == 0:
            messagebox.showinfo(title="Missing", message="No Exams Found\nEnsure Images are PNG, JPG, or JPEG.")
        else:
            folder_path = Folderpath
            Start()

def Destroy():
    for Frame in MainFrame.winfo_children():
        Frame.destroy()

def Template():
    Destroy()
    Screen0 = customtkinter.CTkFrame(MainFrame)
    Screen0.pack(padx=30, pady=30, fill="both", expand=True)

    Screen0.grid_rowconfigure(0, weight=1)
    Screen0.grid_rowconfigure(1, weight=1)
    Screen0.grid_rowconfigure(2, weight=1)
    Screen0.grid_rowconfigure(3, weight=1)
    Screen0.grid_rowconfigure(4, weight=1)
    Screen0.grid_rowconfigure(5, weight=1)
    Screen0.grid_rowconfigure(6, weight=1)
    Screen0.grid_rowconfigure(7, weight=1)
    Screen0.grid_columnconfigure(0, weight=1)

    title = customtkinter.CTkLabel(Screen0, text="DETS Correction System", font=("Arial", 48, "bold"))
    title.grid(row=2, column=0)

    temp10 = customtkinter.CTkImage(dark_image=Image.open(os.path.join(os.path.dirname(__file__), "Data/Templates/Template-10.png")), size=(1558/10, 2068/10))
    temp10_label = customtkinter.CTkLabel(Screen0, image=temp10, text="")
    temp10_label.grid(row=5, column=0, padx=107.1, sticky="w")
    Q10_btn = customtkinter.CTkButton(Screen0, text="10-Questions", font=("Arial", 24, "bold"), command=Temp10, height=60, width=250, corner_radius=10)
    Q10_btn.grid(row=6, column=0, padx=60, sticky="w")

    temp30 = customtkinter.CTkImage(dark_image=Image.open(os.path.join(os.path.dirname(__file__), "Data/Templates/Template-30.png")), size=(1558/10, 2070/10))
    temp30_label = customtkinter.CTkLabel(Screen0, image=temp30, text="")
    temp30_label.grid(row=5, column=0, padx=60)
    Q30_btn = customtkinter.CTkButton(Screen0, text="30-Questions", font=("Arial", 24, "bold"), command=Temp30, height=60, width=250, corner_radius=10)
    Q30_btn.grid(row=6, column=0, padx=60)

    temp60 = customtkinter.CTkImage(dark_image=Image.open(os.path.join(os.path.dirname(__file__), "Data/Templates/Template-60.png")), size=(1558/10, 2070/10))
    temp60_label = customtkinter.CTkLabel(Screen0, image=temp60, text="")
    temp60_label.grid(row=5, column=0, padx=107.1, sticky="e")
    Q60_btn = customtkinter.CTkButton(Screen0, text="60-Questions", font=("Arial", 24, "bold"), command=Temp60, height=60, width=250, corner_radius=10)
    Q60_btn.grid(row=6, column=0, padx=60, sticky="e")

    cash_label = customtkinter.CTkLabel(Screen0, text=f"Cashe: {format_size(get_folder_space(Excel_path)+get_folder_space(result_directory)+get_folder_space(output_directory))}", font=("Arial", 16, "bold"))
    cash_label.grid(row=7, column=0, pady=55, padx=20, sticky="es")
    Clear = customtkinter.CTkButton(Screen0, text="Clear Cache", font=("Arial", 16, "bold"), command=clear_cashe, height=40, width=150, corner_radius=10)
    Clear.grid(row=7, column=0, pady=10, padx=10, sticky="es")

def get_folder_space(folder_path):
    total_size = 0
    with os.scandir(folder_path) as entries:
        for entry in entries:
            if entry.is_file():
                total_size += entry.stat().st_size
            elif entry.is_dir():
                total_size += get_folder_space(entry.path)

    return total_size

def format_size(size_bytes):
    # Convert bytes to a human-readable format
    for unit in ['bytes', 'Kb', 'Mb', 'Gb', 'Tb']:
        if size_bytes < 1024:
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024

def clear_cashe():
    shutil.rmtree(Excel_path)
    os.makedirs(Excel_path)
    shutil.rmtree(result_directory)
    os.makedirs(result_directory)
    shutil.rmtree(output_directory)
    os.makedirs(output_directory)
    Template()

def Temp10():
    global template_image_path
    global num_questions

    num_questions = 10
    template_image_path = "Data/Templates/Template-10.png"

    Start()

def Temp30():
    global template_image_path
    global num_questions

    num_questions = 30
    template_image_path = "Data/Templates/Template-30.png"

    Start()

def Temp60():
    global template_image_path
    global num_questions

    num_questions = 60
    template_image_path = "Data/Templates/Template-60.png"

    Start()

def Start():
    Destroy()
    Screen1 = customtkinter.CTkFrame(MainFrame)
    Screen1.pack(padx=30, pady=30, fill="both", expand=True)

    Screen1.grid_rowconfigure(0, weight=1)
    Screen1.grid_rowconfigure(1, weight=1)
    Screen1.grid_rowconfigure(2, weight=1)
    Screen1.grid_rowconfigure(3, weight=3)
    Screen1.grid_rowconfigure(4, weight=3)
    Screen1.grid_rowconfigure(5, weight=3)
    Screen1.grid_rowconfigure(6, weight=3)
    Screen1.grid_rowconfigure(7, weight=3)
    Screen1.grid_rowconfigure(8, weight=3)
    Screen1.grid_rowconfigure(9, weight=3)
    Screen1.grid_columnconfigure(0, weight=1)


    title = customtkinter.CTkLabel(Screen1, text="DETS Correction System", font=("Arial", 48, "bold"))
    title.grid(row=0, column=0, pady=30)

    Floder_Req = customtkinter.CTkLabel(Screen1, text="Choose The Tests Folder:", font=("Arial", 24, "bold"))
    Floder_Req.grid(row=2, column=0, pady=10)

    Folder_Entry = customtkinter.CTkEntry(Screen1, placeholder_text=folder_path)
    Folder_Entry.grid(row=3, column=0, padx=20, pady=10, sticky="ew")
    Folder_Entry.configure(state="disabled")

    Folder_btn = customtkinter.CTkButton(Screen1, text="Browse", font=("Arial", 16, "bold"), command=OpenFolder, height=40, width=200, corner_radius=10)
    Folder_btn.grid(row=4, column=0, pady=10)
    
    customtkinter.CTkLabel(Screen1, text="").grid(row=5, column=0, padx=20, pady=5)

    File_Req = customtkinter.CTkLabel(Screen1, text="Choose The Student Data Excel File:", font=("Arial", 24, "bold"))
    File_Req.grid(row=6, column=0, pady=10)

    File_Entry = customtkinter.CTkEntry(Screen1, placeholder_text=Students_data_path)
    File_Entry.grid(row=7, column=0, padx=20, pady=10, sticky="ew")
    File_Entry.configure(state="disabled")

    File_btn = customtkinter.CTkButton(Screen1, text="Browse", font=("Arial", 16, "bold"), command=OpenFile, height=40, width=200, corner_radius=10)
    File_btn.grid(row=8, column=0, pady=10)

    Next_btn = customtkinter.CTkButton(Screen1, text="Next", font=("Arial", 24, "bold"), command=Ans, height=50, width=400, corner_radius=10)
    Next_btn.grid(row=9, column=0, pady=60)


def Ans():
    if(folder_path == "Click Browse To Choose Folder..." or Students_data_path == "Click Browse To Choose Excel File..."):
        messagebox.showinfo(title="Missing", message="You didn't Choose Folder or Excel File")
        return
    Destroy()
    Screen2 = customtkinter.CTkFrame(MainFrame)
    Screen2.pack(padx=30, pady=15, fill="x")

    title = customtkinter.CTkLabel(Screen2, text="Choose The Correct Answers", font=("Arial", 48, "bold")).pack(pady=20)

    Screen2_q = customtkinter.CTkScrollableFrame(MainFrame, width=1000, height=500)
    Screen2_q.pack(padx=30, pady=5, expand=True)

    space = customtkinter.CTkLabel(Screen2_q, text="", font=("Arial", 24, "bold")).grid(row=0, column=1, padx=50, pady=10)

    t = customtkinter.CTkLabel(Screen2_q, text="Questions", font=("Arial", 24, "bold")).grid(row=0, column=0, padx=10, pady=10)
    A = customtkinter.CTkLabel(Screen2_q, text="A", font=("Arial", 24, "bold")).grid(row=0, column=2, padx=15, pady=10, sticky="w")
    B = customtkinter.CTkLabel(Screen2_q, text="B", font=("Arial", 24, "bold")).grid(row=0, column=3, padx=15, pady=10, sticky="w")
    C = customtkinter.CTkLabel(Screen2_q, text="C", font=("Arial", 24, "bold")).grid(row=0, column=4, padx=15, pady=10, sticky="w")
    D = customtkinter.CTkLabel(Screen2_q, text="D", font=("Arial", 24, "bold")).grid(row=0, column=5, padx=15, pady=10, sticky="w")
    G = customtkinter.CTkLabel(Screen2_q, text="None", font=("Arial", 24, "bold")).grid(row=0, column=6, padx=15, pady=10, sticky="w")

    question_vars = [customtkinter.StringVar(value="other") for _ in range(num_questions)]

    # Loop through questions
    for i in range(num_questions):
        # Create Label for question number
        q_t = customtkinter.CTkLabel(Screen2_q, text=str(i + 1), font=("Arial", 24, "bold"))
        q_t.grid(row=i + 1, column=0, padx=10, pady=10)

        # Create RadioButtons for options A, B, C, D
        for j, option in enumerate(["A", "B", "C", "D", "G"], start=2):
            q_option = customtkinter.CTkRadioButton(Screen2_q, text="", value=option, variable=question_vars[i])
            q_option.grid(row=i + 1, column=j, padx=15, pady=10)

    
    Screen2_S = customtkinter.CTkFrame(MainFrame, height=500)
    Screen2_S.pack(padx=30, pady=30, fill="both", expand=True)
    
    Next_btn = customtkinter.CTkButton(Screen2_S, text="Submit", font=("Arial", 24, "bold"), command=lambda:Sub(question_vars), height=50, width=170, corner_radius=10)
    Next_btn.pack(expand=True, pady=10)

    Screen2_q.grid_columnconfigure(0, weight=1)

def Sub(question_vars):
    global correcte_answer
    global ANSWER_KEY
    global Excel_name
    global output_path
    
    for i in range(num_questions):
        correcte_answer.extend(question_vars[i].get())
    print(len(correcte_answer))

    if len(correcte_answer) > num_questions:
        correcte_answer = []
        messagebox.showinfo(title="Missing", message="There are Missing Answers")
        return
    
    else:
        if os.path.isdir(output_directory):
            shutil.rmtree(output_directory)
        os.makedirs(output_directory)

        if os.path.isdir(result_directory):
            shutil.rmtree(result_directory)
        os.makedirs(result_directory)
        
        k = 0
        for i in correcte_answer:
            if i == 'A':
                ANSWER_KEY[k] = 0
            elif i == 'B':
                ANSWER_KEY[k] = 1
            elif i == 'C':
                ANSWER_KEY[k] = 2
            elif i == 'D':
                ANSWER_KEY[k] = 3
            elif i == 'G':
                ANSWER_KEY[k] = 6
            k += 1
        print(correcte_answer)
        print(ANSWER_KEY)

        num_exams = 0
        for filename in os.listdir(folder_path):
                if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith('.jpeg'):
                    num_exams += 1

        Destroy()
        Screen3 = customtkinter.CTkFrame(MainFrame)
        Screen3.pack(padx=30, pady=30, fill="both", expand=True)

        title = customtkinter.CTkLabel(Screen3, text="Please Wait...", font=("Arial", 24, "bold"))
        title.place(relx=0.5, rely=0.45, anchor="center")        

        prograssbar = customtkinter.CTkProgressBar(Screen3, orientation="horizontal", height=30, width=450, determinate_speed=(1/num_exams)/0.02)
        prograssbar.pack(expand=True)
        prograssbar.set(0)

        exams = customtkinter.CTkLabel(Screen3, text=f"Finished: 0/{num_exams}", font=("Arial", 24, "bold"))
        exams.place(relx=0.5, rely=0.55, anchor="center")

        threading.Thread(target=start_correct, args=(prograssbar, exams, num_exams,)).start()


def start_correct(prograssbar, exams, num_exams):
    global Excel_name
    global output_path

    start_time = time.time()
    finshed_exams=0
    for filename in os.listdir(folder_path):
        if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith('.jpeg'):
            image_path = os.path.join(folder_path, filename)
            output_path = os.path.join(output_directory, f"aligned_{filename}")
            align_image(template_image_path, image_path, output_path)
            finshed_exams +=1
            exams.configure(text=f"Finished: {finshed_exams}/{num_exams}")
            prograssbar.step()
            
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    formatted_time = format_time(elapsed_time)
    print(f"Time elapsed: {formatted_time}")

    now = datetime.now()
    current_time = now.strftime("%H-%M")
    Excel_name = f"Test_Grades_{current_time}"
    create_excel_sheet(output_directory)

    submit()


def format_time(seconds):
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"


def align_image(template_image_path, input_image_path, output_aligned_image_path):
    img = Image.open(input_image_path)
    img = img.convert('L')
    threshold_img = img.point(lambda p: p > 220 and 255)
    threshold_img.save(output_aligned_image_path)

    template = cv2.imread(template_image_path, 0)
    input_image = cv2.imread(output_aligned_image_path, 0)


    # Initiate the SIFT detector
    sift = cv2.SIFT_create()

    # Find the keypoints and descriptors with SIFT
    kp1, des1 = sift.detectAndCompute(template, None)
    kp2, des2 = sift.detectAndCompute(input_image, None)

    # FLANN parameters
    FLANN_INDEX_KDTREE = 1
    index_params = dict(algorithm=FLANN_INDEX_KDTREE, trees=5)
    search_params = dict(checks=50)

    flann = cv2.FlannBasedMatcher(index_params, search_params)
    matches = flann.knnMatch(des1, des2, k=2)

    # Ratio test to filter good matches
    good_matches = []
    for m, n in matches:
        if m.distance < 0.7 * n.distance:
            good_matches.append(m)

    if len(good_matches) > 4:
        src_pts = np.float32([kp1[m.queryIdx].pt for m in good_matches]).reshape(-1, 1, 2)
        dst_pts = np.float32([kp2[m.trainIdx].pt for m in good_matches]).reshape(-1, 1, 2)

        # Calculate the transformation matrix
        M, _ = cv2.findHomography(dst_pts, src_pts, cv2.RANSAC, 5.0)

        # Apply the transformation to the input image
        result = cv2.warpPerspective(input_image, M, (template.shape[1], template.shape[0]))

        # Save the aligned image
        cv2.imwrite(output_aligned_image_path, result)

        print(f"Aligned image saved at {output_aligned_image_path}")
    else:
        print("Not enough matches are found - %d/%d" % (len(good_matches), 4))

    Scan(output_aligned_image_path)


def encode(serial_number):
    test_code = ""
    for char in serial_number:
        test_code += str(ord(char))
    return test_code
    
def extract_serial_number(image):
    # Convert the image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Apply image preprocessing techniques
    processed_image = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 85, 11)

    # Use pytesseract to extract text
    extracted_text = pytesseract.image_to_string(processed_image)

    # Search for the serial number pattern
    serial_number = None
    for line in extracted_text.split('\n'):
        if 'S/N: ' in line:
            serial_number = line.split('S/N: ')[1]
            break

    return serial_number

def is_shaded(img, x, y, radius, threshold):
    height, width, _ = img.shape
    total_color = np.array([0, 0, 0], dtype=np.uint64)
    count = 0

    for j in range(-radius, radius + 1):
        for i in range(-radius, radius + 1):
            if 0 <= y + j < height and 0 <= x + i < width:
                total_color += img[y + j, x + i]
                count += 1

    if count > 0:
        average_color = total_color / count
        return sum(average_color) < threshold * 3
    else:
        return False

def Scan(image_path):
    # Load the image
    img = cv2.imread(image_path)

    # Draw a point on the image at the specified coordinates
    color = (0, 0, 255)  # Red color in BGR
    thickness = -1  # Fill the circle
    radius_circle = 10  # Radius of the circle
    
    Student_Answers = {} # Answers of students
    
    if num_questions == 10:
        # Define the coordinates
        x, y = 306, 653

        # Define the radius and shading threshold
        radius = 20
        shaded_threshold = 150
        for j in range(num_questions):
            shaded = 0
            for i in range(4):
                # Calculate the coordinates
                point_x = x + 120 * i
                point_y = y + 120 * j

                # Check if the calculated coordinates are within the image dimensions
                if point_x < img.shape[1] and point_y < img.shape[0]:
                    # Check if the area is shaded
                    if is_shaded(img, point_x, point_y, radius, shaded_threshold):
                        Student_Answers[j] = i
                        cv2.circle(img, (point_x, point_y), radius_circle, color, thickness)
                        shaded += 1
                    
                    if shaded > 1:
                        Student_Answers[j] = 4
                    
                    if i == 3 and j not in Student_Answers:
                        Student_Answers[j] = 5
    elif num_questions == 30:
        # Define the coordinates
        x, y = 353, 614

        # Define the radius and shading threshold
        radius = 20
        shaded_threshold = 170
        for j in range(num_questions):
            shaded = 0
            if j == 15:
                x, y = 1022, 614
            for i in range(4):
                # Calculate the coordinates
                point_x = x + 94 * i
                point_y = y + 85 * j
                if j > 14:
                    point_y = y + 85 * (j-15)

                # Check if the calculated coordinates are within the image dimensions
                if point_x < img.shape[1] and point_y < img.shape[0]:
                    # Check if the area is shaded
                    if is_shaded(img, point_x, point_y, radius, shaded_threshold):
                        Student_Answers[j] = i
                        cv2.circle(img, (point_x, point_y), radius_circle, color, thickness)
                        shaded += 1
                    
                    if shaded > 1:
                        Student_Answers[j] = 4
                    
                    if i == 3 and j not in Student_Answers:
                        Student_Answers[j] = 5
    elif num_questions == 60:
        # Define the coordinates
        x, y = 242, 620

        # Define the radius and shading threshold
        radius = 20
        shaded_threshold = 170
        for j in range(num_questions):
            shaded = 0
            if j == 20:
                x, y = 696, 620
            elif j == 40:
                x, y = 1159, 620
            for i in range(4):
                # Calculate the coordinates
                point_x = x + 84 * i
                point_y = y + 63 * j
                if j > 19 and j < 40:
                    point_y = y + 63 * (j-20)
                elif j > 39:
                    point_y = y + 63 * (j-40)

                # Check if the calculated coordinates are within the image dimensions
                if point_x < img.shape[1] and point_y < img.shape[0]:
                    # Check if the area is shaded
                    if is_shaded(img, point_x, point_y, radius, shaded_threshold):
                        Student_Answers[j] = i
                        cv2.circle(img, (point_x, point_y), radius_circle, color, thickness)
                        shaded += 1
                    
                    if shaded > 1:
                        Student_Answers[j] = 4
                    
                    if i == 3 and j not in Student_Answers:
                        Student_Answers[j] = 5

    Correction(Student_Answers, img)

def Correction(Student_Answers, image):
    #The number of correct answers for the student
    correct = 0
    
    #Answer List
    Answer_list = []
    
    for i in range(num_questions):
        if ANSWER_KEY[i] == Student_Answers[i]:
            correct += 1
        
        if Student_Answers[i] == 0:
            Answer_list.extend('A')
        elif Student_Answers[i] == 1:
            Answer_list.extend('B')
        elif Student_Answers[i] == 2:
            Answer_list.extend('C')
        elif Student_Answers[i] == 3:
            Answer_list.extend('D')
        elif Student_Answers[i] == 4:
            Answer_list.extend('M')
        elif Student_Answers[i] == 5:
            Answer_list.extend('N')
            
    serial_number = extract_serial_number(image)
    serial_numbers[os.path.basename(output_path)] = serial_number
    
    test_code = encode(serial_number)
    test_codes[serial_number] = test_code
    
    test_serial_codes[test_code] = serial_number
    
    # Draw a square on the image
    x, y, w, h = 122, 421, 339, 68
    cv2.rectangle(image, (x, y), (x + w, y + h), (0, 0, 0), -1)
    cv2.putText(image, test_code, (x, y + 115), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 0, 0), 2)
    
    # Save the aligned image with points
    result_path = f'{result_directory}/{test_code}.png'
    cv2.imwrite(result_path, image)
    corrected_tests[test_code] = {'answers': Answer_list, 'grade': correct, 'link': result_path}
    #print(corrected_tests[test_code])
    cv2.waitKey(0)
    
def create_excel_sheet(folder_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Results"
    green_count_formula = "0"
    
    # Writing headers
    sheet.cell(row=1, column=1, value="Test Code")
    questions = len(ANSWER_KEY)
    for i in range(questions):
        question_num = 'Question ' + str(i + 1)
        sheet.cell(row=1, column=i + 2, value=question_num)
    sheet.cell(row=1, column=questions + 2, value="Final Grade")
    sheet.cell(row=1, column=questions + 3, value="Link to Corrected Test")
    
    # Writing data
    row_num = 2
    for filename in os.listdir(folder_path):
        if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith('.jpeg'):
            image_path = os.path.join(folder_path, filename)
            test_name = os.path.basename(image_path)
            serial_number = serial_numbers[test_name]
            test_code = test_codes[serial_number]
            sheet.cell(row=row_num, column=1, value=test_code)
            for i, answer in enumerate(corrected_tests[test_code]['answers']):
                sheet.cell(row=row_num, column=i + 2, value=answer)
                counter = f'COUNTIF({get_column_letter(i + 2)}{row_num}, "{correcte_answer[i]}")'
                if i > 0:
                    green_count_formula += f"+{counter}"   
                else:
                    green_count_formula = f"={counter}"
            
            # Marking cells based on correctness
            grade_cell = f'{get_column_letter(questions + 2)}{row_num}'
            sheet[grade_cell] = green_count_formula
            
            # Adding hyperlink to the cell
            hyperlink_cell = sheet.cell(row=row_num, column=questions + 3)
            hyperlink_cell.value = f"Link to Corrected Test {test_code}"
            hyperlink_cell.font = Font(underline='single', color="0563C1")
            hyperlink = f"file:///{os.path.abspath(result_directory)}/{test_code}.png"
            sheet.cell(row=row_num, column=questions + 3).hyperlink = hyperlink
            sheet.cell(row=row_num, column=questions + 3).style = "Hyperlink"
            row_num += 1
        
    # Color the cells according to the answer
    for row in range(2, row_num):
        for col in range(2, questions + 2):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value == correcte_answer[col - 2]:
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Good theme
            elif cell_value == "N" or cell_value == "M":
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="FBFF7A", end_color="FBFF7A", fill_type="solid")  # Manually theme
            else:
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")  # Bad theme


    # Adding conditional formatting
    for col in range(2, questions + 2):
        rule = CellIsRule(operator="equal", formula=[f'"{correcte_answer[col - 2]}"'], stopIfTrue=True, fill=PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"))
        sheet.conditional_formatting.add(f'{sheet.cell(2, col).column_letter}2:{sheet.cell(row_num - 1, col).column_letter}{row_num - 1}', rule)

        rule = CellIsRule(operator="equal", formula=['"N"'], stopIfTrue=True, fill=PatternFill(start_color="FBFF7A", end_color="FBFF7A", fill_type="solid"))
        sheet.conditional_formatting.add(f'{sheet.cell(2, col).column_letter}2:{sheet.cell(row_num - 1, col).column_letter}{row_num - 1}', rule)

        rule = CellIsRule(operator="equal", formula=['"M"'], stopIfTrue=True, fill=PatternFill(start_color="FBFF7A", end_color="FBFF7A", fill_type="solid"))
        sheet.conditional_formatting.add(f'{sheet.cell(2, col).column_letter}2:{sheet.cell(row_num - 1, col).column_letter}{row_num - 1}', rule)

        rule = CellIsRule(operator="notEqual", formula=[f'"{correcte_answer[col - 2]}"', '"N"', '"M"'], stopIfTrue=True, fill=PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"))
        sheet.conditional_formatting.add(f'{sheet.cell(2, col).column_letter}2:{sheet.cell(row_num - 1, col).column_letter}{row_num - 1}', rule)

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        sheet.column_dimensions[f'{get_column_letter(questions + 2)}'].width = 11
            
    excel_file_path = os.path.join(Excel_path, f"{Excel_name}.xlsx")
    wb.save(excel_file_path)
    print(f"Excel file saved at {excel_file_path}")
    # Open the Excel file automatically
    os.startfile(excel_file_path)
    
def create_final_grades_excel(excel_file_path, test_serial_codes, Students_data_path):
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheet = wb.active

    final_grades_wb = openpyxl.Workbook()
    final_grades_sheet = final_grades_wb.active
    final_grades_sheet.title = "Final Grades"
    questions = len(ANSWER_KEY)
    
    # Writing headers
    final_grades_sheet.cell(row=1, column=1, value="Serial Number")
    final_grades_sheet.cell(row=1, column=2, value="Name")
    final_grades_sheet.cell(row=1, column=3, value="Acadmic Number")
    final_grades_sheet.cell(row=1, column=4, value="Final Grade")

    # Loading student data
    student_data_wb = openpyxl.load_workbook(Students_data_path)
    student_sheet = student_data_wb.active

    student_data = {}
    for row in student_sheet.iter_rows(min_row=2, max_row=student_sheet.max_row, max_col=3):
        serial_number = row[0].value
        name = row[1].value
        acadmic_number = row[2].value
        student_data[serial_number] = {'name': name, 'acadmic_number': acadmic_number}

    row_num = 2
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        test_code = row[0].value
        serial_number = test_serial_codes[test_code]
        final_grade = row[questions + 1].value
        if serial_number in student_data:
            name = student_data[serial_number]['name']
            acadmic_number = student_data[serial_number]['acadmic_number']
            final_grades_sheet.cell(row=row_num, column=1, value=serial_number)
            final_grades_sheet.cell(row=row_num, column=2, value=name)
            final_grades_sheet.cell(row=row_num, column=3, value=acadmic_number)
            final_grades_sheet.cell(row=row_num, column=4, value=final_grade)
            row_num += 1
        else:
            print(f"Serial number {serial_number} not found in the student data.")

    for column_cells in final_grades_sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        final_grades_sheet.column_dimensions[column_cells[0].column_letter].width = length + 2            
    
    final_grades_wb.save(final_grades_file_path)
    
    # Protecting the sheet
    final_grades_wb = openpyxl.load_workbook(final_grades_file_path)
    final_grades_sheet = final_grades_wb.active
    final_grades_sheet.protection.sheet = True

    # Set a password for the sheet (optional)
    final_grades_sheet.protection.password = "securepassword"
    
    final_grades_wb.save(final_grades_file_path)
    print(f"Final grades Excel file saved at {final_grades_file_path}")    

def submit():
    Destroy()
    Screen4 = customtkinter.CTkFrame(MainFrame)
    Screen4.pack(padx=30, pady=30, fill="both", expand=True)
    title = customtkinter.CTkLabel(Screen4, text="Did You Check the Grades?!", font=("Arial", 48, "bold"))
    title.place(relx=0.5, rely=0.2, anchor="center")
    Open_btn = customtkinter.CTkButton(Screen4, text="Open", font=("Arial", 24, "bold"), command=OpenExcel, height=60, width=300, corner_radius=10)
    Open_btn.place(relx=0.5, rely=0.5, anchor="center")
    Done_btn = customtkinter.CTkButton(Screen4, text="Done", font=("Arial", 24, "bold"), command=done, height=60, width=300, corner_radius=10)
    Done_btn.place(relx=0.5, rely=0.61, anchor="center")
    
def done():
    global final_grades_file_path
    x = filedialog.asksaveasfile(defaultextension=".xlsx", filetypes=[("Excel File","*.xlsx")], initialfile="Final Grades.xlsx")
    if x is None:
        return
    final_grades_file_path = x.name
    Destroy()
    create_final_grades_excel(os.path.join(Excel_path, f"{Excel_name}.xlsx"), test_serial_codes, Students_data_path)
    Screen5 = customtkinter.CTkFrame(MainFrame)
    Screen5.pack(padx=30, pady=30, fill="both", expand=True)
    title = customtkinter.CTkLabel(Screen5, text="Thanks For Using\nDETS Correction System", font=("Arial", 48, "bold"))
    title.place(relx=0.5, rely=0.2, anchor="center")
    Open_btn = customtkinter.CTkButton(Screen5, text="Open", font=("Arial", 24, "bold"), command=OpenFinalExcel, height=60, width=300, corner_radius=10)
    Open_btn.place(relx=0.5, rely=0.5, anchor="center")
    Done_btn = customtkinter.CTkButton(Screen5, text="Exit", font=("Arial", 24, "bold"), command=lambda: sys.exit(), height=60, width=300, corner_radius=10)
    Done_btn.place(relx=0.5, rely=0.61, anchor="center")

def OpenExcel():
    os.startfile(os.path.join(Excel_path, f"{Excel_name}.xlsx"))

def OpenFinalExcel():
    os.startfile(final_grades_file_path)

version = customtkinter.CTkLabel(app, text="version 1.2", bg_color="#212121", font=("Arial", 12, "bold"))
version.place(relx=0.01, rely=1, anchor="sw")

Main.grid_rowconfigure(0, weight=1)
Main.grid_rowconfigure(1, weight=1)
Main.grid_rowconfigure(2, weight=1)
Main.grid_rowconfigure(3, weight=1)
Main.grid_rowconfigure(4, weight=1)
Main.grid_rowconfigure(5, weight=1)
Main.grid_rowconfigure(6, weight=1)
Main.grid_rowconfigure(7, weight=1)
Main.grid_rowconfigure(8, weight=1)
Main.grid_rowconfigure(9, weight=1)
Main.grid_rowconfigure(10, weight=1)
Main.grid_columnconfigure(0, weight=1)


title = customtkinter.CTkLabel(Main, text="DETS Correction System", font=("Arial", 48, "bold"))
title.grid(row=2, column=0)

Start_btn = customtkinter.CTkButton(Main, text="Start", font=("Arial", 24, "bold"), command=Template, height=60, width=300, corner_radius=10)
Start_btn.grid(row=5, column=0)
Exit_btn = customtkinter.CTkButton(Main, text="Exit", font=("Arial", 24, "bold"), command=lambda: sys.exit(), height=60, width=300, corner_radius=10)
Exit_btn.grid(row=6, column=0)


if os.path.isdir(Excel_path):
    pass
else:
    os.makedirs(Excel_path)

app.mainloop()